import streamlit as st
import pandas as pd
import io
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.database.connection import connect_to_gsheets


def generate_excel_report(report_df: pd.DataFrame, sheet_name="Marks_Report") -> bytes:
    """Generates an openpyxl-styled Excel workbook as a downloadable binary buffer."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(report_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = max(
                report_df[col_name].astype(str).map(len).max() if not report_df.empty else 0,
                len(str(col_name))
            ) + 4
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len, 12)

    output.seek(0)
    return output.getvalue()


def save_marks_to_gsheets(edited_df: pd.DataFrame, exam_name: str, subject: str, db: dict) -> int:
    """Formats edited dataframe and appends valid records to Google Sheets Marks_Log."""
    client = connect_to_gsheets()
    sheet = client.open("PS Cadet College - Master Examination Database").worksheet("Marks_Log")

    # Match Exam_ID from exam_scheme or Grading_System
    exam_scheme = db.get("exam_scheme", pd.DataFrame())
    grading_df = db.get("Grading_System", pd.DataFrame())

    exam_id = exam_name
    if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns and "Exam_ID" in exam_scheme.columns:
        match = exam_scheme.loc[exam_scheme["Exam_Name"] == exam_name, "Exam_ID"]
        if not match.empty:
            exam_id = match.values[0]
    elif not grading_df.empty and "Exam_Name" in grading_df.columns and "Exam_ID" in grading_df.columns:
        match = grading_df.loc[grading_df["Exam_Name"] == exam_name, "Exam_ID"]
        if not match.empty:
            exam_id = match.values[0]

    records_to_add = []
    assert isinstance(edited_df, pd.DataFrame)
    id_col = "Kit_No" if "Kit_No" in edited_df.columns else "Student_ID"

    absent_keywords = {"ab", "a", "absent", "a/b", "n/a", "na", "-"}

    for _, row in edited_df.iterrows():
        raw = row['Marks_Obtained']
        marks_val = str(raw).strip().lower()
        if marks_val == "" or marks_val == "nan" or pd.isna(raw):
            continue

        if marks_val in absent_keywords:
            canonical = "Absent"
        else:
            try:
                float(marks_val)
                canonical = str(raw).strip()
            except ValueError:
                canonical = "Absent"

        submission_id = str(uuid.uuid4())[:8]
        student_code = str(row[id_col]).strip()
        records_to_add.append([
            submission_id,
            student_code,
            str(exam_id).strip(),
            str(subject).strip(),
            canonical
        ])

    if records_to_add:
        sheet.append_rows(records_to_add)
        st.cache_data.clear()  # Purge cache to reflect updates instantly
        return len(records_to_add)
    return 0