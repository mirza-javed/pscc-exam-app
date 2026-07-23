import streamlit as st
import pandas as pd
import gspread
import uuid
import io
import textwrap
import matplotlib.pyplot as plt
import seaborn as sns
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="PS Cadet College Karachi Exam Portal",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Define API Scopes for Google Drive & Google Sheets
SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

# --- ADVANCED DESIGN SYSTEM & RESPONSIVE CSS INJECTION ---
st.markdown("""
<style>
    /* Google Font Import */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Outfit:wght@400;500;600;700&display=swap');
    
    /* Root Variables & Colors */
    :root {
        --primary-gradient: linear-gradient(135deg, #0f172a 0%, #1e3a8a 50%, #2563eb 100%);
        --card-bg: #ffffff;
        --card-border: #e2e8f0;
        --accent-blue: #2563eb;
        --text-dark: #0f172a;
        --text-muted: #64748b;
    }
    
    /* Global App Styling */
    .stApp {
        font-family: 'Inter', sans-serif;
        background-color: #f8fafc;
        color: var(--text-dark);
    }
    
    h1, h2, h3, h4, h5 {
        font-family: 'Outfit', 'Inter', sans-serif;
    }

    /* Main Hero Banner */
    .hero-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 40%, #2563eb 100%);
        color: white;
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 10px 25px -5px rgba(37, 99, 235, 0.25);
        position: relative;
        overflow: hidden;
    }
    .hero-header::after {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: rgba(255, 255, 255, 0.05);
        border-radius: 50%;
        pointer-events: none;
    }
    .hero-header h1 {
        margin: 0;
        font-size: 2.2rem;
        font-weight: 800;
        color: #ffffff !important;
        letter-spacing: -0.02em;
    }
    .hero-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.95;
        font-size: 1.05rem;
        font-weight: 400;
    }

    /* Custom Metric Stat Cards */
    .stat-card {
        background: #ffffff;
        border: 1px solid var(--card-border);
        border-radius: 14px;
        padding: 1.25rem 1.5rem;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03);
        transition: all 0.25s ease-in-out;
        border-top: 4px solid var(--accent-blue);
    }
    .stat-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 20px rgba(37, 99, 235, 0.12);
    }
    .stat-label {
        font-size: 0.85rem;
        font-weight: 600;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 0.3rem;
    }
    .stat-value {
        font-size: 1.8rem;
        font-weight: 800;
        color: var(--text-dark);
        font-family: 'Outfit', sans-serif;
    }
    .stat-badge {
        display: inline-block;
        margin-top: 0.4rem;
        padding: 0.2rem 0.6rem;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 700;
    }

    /* Grade Status Badges */
    .badge-aplus { background-color: #dcfce7; color: #15803d; border: 1px solid #bbf7d0; }
    .badge-a     { background-color: #e0f2fe; color: #0369a1; border: 1px solid #bae6fd; }
    .badge-b     { background-color: #fef9c3; color: #a16207; border: 1px solid #fef08a; }
    .badge-c     { background-color: #ffedd5; color: #c2410c; border: 1px solid #fed7aa; }
    .badge-d     { background-color: #f3e8ff; color: #7e22ce; border: 1px solid #e9d5ff; }
    .badge-f     { background-color: #fee2e2; color: #b91c1c; border: 1px solid #fecaca; }

    /* Touch Friendly Form Inputs & Selectboxes */
    .stSelectbox div[data-baseweb="select"] > div {
        border-radius: 10px !important;
        min-height: 44px !important;
        border-color: #cbd5e1 !important;
    }
    .stButton > button {
        border-radius: 10px !important;
        min-height: 46px !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        transition: all 0.2s ease-in-out !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(37, 99, 235, 0.2) !important;
    }

    /* Custom Streamlit Tabs Design */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: #f1f5f9;
        padding: 6px;
        border-radius: 12px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 44px;
        white-space: pre;
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.92rem;
        color: #475569;
        transition: all 0.2s ease;
    }
    .stTabs [aria-selected="true"] {
        background-color: #ffffff !important;
        color: #1e3a8a !important;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
    }

    /* Mobile Responsive Optimizations */
    @media (max-width: 768px) {
        .hero-header {
            padding: 1.5rem 1.25rem;
            border-radius: 12px;
        }
        .hero-header h1 {
            font-size: 1.5rem !important;
        }
        .hero-header p {
            font-size: 0.9rem !important;
        }
        .stat-value {
            font-size: 1.4rem !important;
        }
        /* Mobile horizontal scroll for dataframes */
        .stDataFrame, .stDataEditor {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch;
        }
        /* Touch target padding on phone */
        .stButton > button {
            width: 100% !important;
        }
    }
</style>
""", unsafe_allow_html=True)


# --- GOOGLE SHEETS CONNECTION & DATA LOADING ---
@st.cache_resource
def connect_to_gsheets():
    """Authenticates using the secrets.toml file or Streamlit Cloud Secrets."""
    if "gcp_service_account" not in st.secrets:
        st.error("🔑 **GCP Service Account Credentials Missing!**")
        st.info(
            "### How to fix this on Streamlit Community Cloud:\n\n"
            "1. Open your app dashboard at **[share.streamlit.io](https://share.streamlit.io/)**.\n"
            "2. Click the **`⋮` (Options)** or ⚙️ **Settings** icon next to your app.\n"
            "3. Click on the **Secrets** tab on the left.\n"
            "4. Paste your `[gcp_service_account]` section from your local `.streamlit/secrets.toml` file into the editor.\n"
            "5. Click **Save**. The app will automatically reboot and connect."
        )
        st.stop()

    try:
        creds_dict = st.secrets["gcp_service_account"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, SCOPE)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        st.error(f"❌ **Failed to authenticate with Google Sheets API:** {e}")
        st.stop()


@st.cache_data(ttl=600)  # Refresh cache every 10 minutes
def load_database():
    """Pulls all database tabs into a dictionary of Pandas DataFrames with clean header normalization."""
    client = connect_to_gsheets()
    sheet = client.open("PS Cadet College - Master Examination Database")

    def fetch_tab(tab_name):
        try:
            data = sheet.worksheet(tab_name).get_all_values()
            if not data:
                return pd.DataFrame()

            headers = [str(h).strip() for h in data[0]]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=headers)

            # Clean string values across all cells (trim spaces)
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()

            # Normalize Kit_No vs Student_ID & Group vs Stream for Students
            if tab_name == "Students":
                if "Kit_No" in df.columns and "Student_ID" not in df.columns:
                    df["Student_ID"] = df["Kit_No"]
                elif "Student_ID" in df.columns and "Kit_No" not in df.columns:
                    df["Kit_No"] = df["Student_ID"]

                if "Group" in df.columns and "Stream" not in df.columns:
                    df["Stream"] = df["Group"]
                elif "Stream" in df.columns and "Group" not in df.columns:
                    df["Group"] = df["Stream"]

                if "Name" not in df.columns:
                    for alt in ["Full_Name", "Full Name", "Student_Name", "Student Name"]:
                        if alt in df.columns:
                            df["Name"] = df[alt]
                            break

            # Normalize Kit_No vs Student_ID for Marks_Log
            if tab_name == "Marks_Log":
                if "Kit_No" in df.columns and "Student_ID" not in df.columns:
                    df["Student_ID"] = df["Kit_No"]
                elif "Student_ID" in df.columns and "Kit_No" not in df.columns:
                    df["Kit_No"] = df["Student_ID"]

            # Normalize Name and Full_Name for Staff_Directory
            if tab_name == "Staff_Directory":
                if "Name" not in df.columns and "Full_Name" in df.columns:
                    df["Name"] = df["Full_Name"]
                elif "Full_Name" not in df.columns and "Name" in df.columns:
                    df["Full_Name"] = df["Name"]

            return df
        except Exception:
            return pd.DataFrame()

    db = {
        "Students": fetch_tab("Students"),
        "Staff_Directory": fetch_tab("Staff_Directory"),
        "Teaching_Assignments": fetch_tab("Teaching_Assignments"),
        "Grading_System": fetch_tab("Grading_System"),
        "exam_scheme": fetch_tab("exam_scheme"),
        "Marks_Log": fetch_tab("Marks_Log"),
        "Group_Subjects": fetch_tab("Group_Subjects"),
        "Subjects_Master": fetch_tab("Subjects_Master")
    }
    return db


def get_all_available_subjects(db: dict) -> list:
    """Collects all unique academic subjects across Group_Subjects, exam_scheme, Subjects_Master, and Teaching_Assignments."""
    subjects = set()

    # 1. Group_Subjects tab
    gs_df = db.get("Group_Subjects", pd.DataFrame())
    if not gs_df.empty:
        for col in gs_df.columns:
            for val in gs_df[col].dropna():
                s = str(val).strip()
                if s and not s.startswith("Subjects_of_"):
                    subjects.add(s)

    # 2. exam_scheme tab
    es_df = db.get("exam_scheme", pd.DataFrame())
    if not es_df.empty and "Subject" in es_df.columns:
        for val in es_df["Subject"].dropna():
            s = str(val).strip()
            if s:
                subjects.add(s)

    # 3. Subjects_Master tab
    sm_df = db.get("Subjects_Master", pd.DataFrame())
    if not sm_df.empty:
        col = "Subject_Name" if "Subject_Name" in sm_df.columns else ("Subject" if "Subject" in sm_df.columns else None)
        if col:
            for val in sm_df[col].dropna():
                s = str(val).strip()
                if s:
                    subjects.add(s)

    # 4. Teaching_Assignments tab
    ta_df = db.get("Teaching_Assignments", pd.DataFrame())
    if not ta_df.empty and "Subject" in ta_df.columns:
        for val in ta_df["Subject"].dropna():
            s = str(val).strip()
            if s:
                subjects.add(s)

    return sorted(list(subjects)) if subjects else ["English", "Urdu", "Maths", "Physics", "Chemistry", "Islamiat", "Biology", "Computer Science", "Pakistan Studies", "Sindhi", "Manners"]


# --- HELPER LOGIC: GRADE THRESHOLD MAPPING ---
def calculate_grade_info(pct: float, grading_df: pd.DataFrame = None) -> dict:
    """
    Computes Letter Grade, Remarks, and Pass/Fail status based on Percentage.
    Leverages Grading_System threshold mappings if defined, or falls back to standard thresholds.
    """
    pct = round(float(pct), 2)

    # Check if threshold columns exist in Grading_System
    if grading_df is not None and not grading_df.empty:
        min_col = "Min_Percentage" if "Min_Percentage" in grading_df.columns else ("Min Percentage" if "Min Percentage" in grading_df.columns else None)
        max_col = "Max_Percentage" if "Max_Percentage" in grading_df.columns else ("Max Percentage" if "Max Percentage" in grading_df.columns else None)

        if min_col and max_col:
            g_df = grading_df.copy()
            g_df["Min_Pct"] = pd.to_numeric(g_df[min_col].astype(str).str.replace("%", ""), errors="coerce")
            g_df["Max_Pct"] = pd.to_numeric(g_df[max_col].astype(str).str.replace("%", ""), errors="coerce")
            g_df = g_df.dropna(subset=["Min_Pct", "Max_Pct"])

            for _, row in g_df.iterrows():
                if row["Min_Pct"] <= pct <= row["Max_Pct"]:
                    grade = str(row.get("Grade", "")).strip() or "N/A"
                    remarks = str(row.get("Remarks", "")).strip() or "Satisfactory"
                    status = "PASS" if grade not in ["F", "Fail", "FAIL", "U"] and pct >= 40 else "FAIL"
                    return {"grade": grade, "remarks": remarks, "status": status}

    # Standard Fallback Thresholds (PSCC Academic Standard)
    if pct >= 95:
        return {"grade": "A++", "remarks": "Exceptional", "status": "PASS"}
    elif pct >= 90:
        return {"grade": "A+", "remarks": "Outstanding", "status": "PASS"}
    elif pct >= 85:
        return {"grade": "A", "remarks": "Excellent", "status": "PASS"}
    elif pct >= 80:
        return {"grade": "B++", "remarks": "Very Good", "status": "PASS"}
    elif pct >= 75:
        return {"grade": "B+", "remarks": "Good", "status": "PASS"}
    elif pct >= 70:
        return {"grade": "B", "remarks": "Fairly Good", "status": "PASS"}
    elif pct >= 60:
        return {"grade": "C", "remarks": "Above Average", "status": "PASS"}
    elif pct >= 50:
        return {"grade": "D", "remarks": "Average", "status": "PASS"}
    elif pct >= 40:
        return {"grade": "E", "remarks": "Below Average", "status": "PASS"}
    else:
        return {"grade": "U", "remarks": "Fail / Unsatisfactory", "status": "FAIL"}


# --- HELPER LOGIC: STAFF ROLE & PERMISSIONS ---
def get_staff_permissions(user_info: dict, db: dict):
    """
    Evaluates permissions for logged-in staff member based on role and assignments:
    1. Global Access: Principal, V. Principal, Section_Head, Admin_Exam, In-charge_Exam.
    2. Class_Teacher: Full access to assigned class (Class_Teacher_Of) & section (Section_Of) for ALL subjects.
    3. Teacher (Subject Teacher): Access restricted to assigned subjects in Teaching_Assignments.
    """
    user_info = user_info or {}
    role = str(user_info.get('Role', '')).strip().lower()
    responsibility = str(user_info.get('Responsibility', '')).strip().lower()
    teacher_id = str(user_info.get('Teacher_ID', '')).strip()

    admin_keywords = [
        'principal', 'v. principal', 'v_principal', 'vice principal',
        'section_head', 'section head', 'admin_exam', 'admin exam',
        'in-charge_exam', 'in-charge examination', 'examination incharge',
        'incharge examination', 'in-charge exam', 'admin', 'administrator'
    ]
    is_admin = any(k in role or k in responsibility for k in admin_keywords)

    if is_admin:
        return {
            "is_admin": True,
            "is_class_teacher": False,
            "assigned_grades": [],
            "assigned_sections": {},
            "assigned_subjects": {}
        }

    assigned_grades_set = set()
    assigned_sections_map = {}   # grade -> set of sections
    assigned_subjects_map = {}   # (grade, section) -> set of subjects
    class_teacher_scopes = set() # (grade, section) where user is class teacher

    # Check Class_Teacher_Of & Section_Of in user_info or Staff_Directory
    class_teacher_of = str(user_info.get("Class_Teacher_Of", user_info.get("Class_Incharge_Of", ""))).strip()
    section_of = str(user_info.get("Section_Of", "")).strip()

    if (class_teacher_of and class_teacher_of.lower() not in ["none", "", "nan"] and
        section_of and section_of.lower() not in ["none", "", "nan"]):
        c_grade = class_teacher_of
        c_sec = section_of
        assigned_grades_set.add(c_grade)
        assigned_sections_map.setdefault(c_grade, set()).add(c_sec)
        class_teacher_scopes.add((c_grade, c_sec))

    # Parse Teaching_Assignments tab
    assignments_df = db.get("Teaching_Assignments", pd.DataFrame())
    if not assignments_df.empty and teacher_id:
        user_assignments = assignments_df[assignments_df["Teacher_ID"].astype(str).str.strip() == teacher_id].copy()

        grade_col = "Assigned_Grade" if "Assigned_Grade" in user_assignments.columns else "Grade"
        subject_col = "Subject" if "Subject" in user_assignments.columns else "Subject_Name"

        section_flag_cols = [
            col for col in user_assignments.columns
            if col.startswith("Assigned_Section_") or (col.startswith("Section_") and col != "Section_Name")
        ]

        truthy_values = {"1", "true", "yes", "y", "t"}

        if section_flag_cols:
            for _, row in user_assignments.iterrows():
                grade = str(row.get(grade_col, "")).strip()
                subject = str(row.get(subject_col, "")).strip()
                if not grade or not subject:
                    continue

                for flag_col in section_flag_cols:
                    val = str(row.get(flag_col, "")).strip().lower()
                    if val in truthy_values:
                        if flag_col.startswith("Assigned_Section_"):
                            sec_name = flag_col[len("Assigned_Section_"):].strip()
                        else:
                            sec_name = flag_col[len("Section_"):].strip()

                        if sec_name:
                            assigned_grades_set.add(grade)
                            assigned_sections_map.setdefault(grade, set()).add(sec_name)
                            assigned_subjects_map.setdefault((grade, sec_name), set()).add(subject)
        else:
            section_col = "Assigned_Section" if "Assigned_Section" in user_assignments.columns else "Section"
            for _, row in user_assignments.iterrows():
                grade = str(row.get(grade_col, "")).strip()
                section = str(row.get(section_col, "")).strip()
                subject = str(row.get(subject_col, "")).strip()
                if grade and section and subject:
                    assigned_grades_set.add(grade)
                    assigned_sections_map.setdefault(grade, set()).add(section)
                    assigned_subjects_map.setdefault((grade, section), set()).add(subject)

    # Class Teachers get FULL access to ALL subjects for their assigned class & section
    all_subjects_list = get_all_available_subjects(db)
    for (c_grade, c_sec) in class_teacher_scopes:
        assigned_subjects_map[(c_grade, c_sec)] = set(all_subjects_list)

    def sort_key(val):
        val_str = str(val)
        return (0, int(val_str)) if val_str.isdigit() else (1, val_str)

    assigned_grades = sorted(list(assigned_grades_set), key=sort_key)
    assigned_sections = {g: sorted(list(secs)) for g, secs in assigned_sections_map.items()}
    assigned_subjects = {k: sorted(list(subs)) for k, subs in assigned_subjects_map.items()}

    return {
        "is_admin": False,
        "is_class_teacher": len(class_teacher_scopes) > 0,
        "class_teacher_scopes": class_teacher_scopes,
        "assigned_grades": assigned_grades,
        "assigned_sections": assigned_sections,
        "assigned_subjects": assigned_subjects
    }


# --- HELPER LOGIC: STYLED EXCEL EXPORT ---
def generate_excel_report(report_df: pd.DataFrame, sheet_name="Marks_Report") -> bytes:
    """Generates an openpyxl-styled Excel workbook as a downloadable binary buffer."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(report_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = max(
                report_df[col_name].astype(str).map(len).max() if not report_df.empty else 0,
                len(str(col_name))
            ) + 4
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len, 12)

    output.seek(0)
    return output.getvalue()


# --- SESSION STATE INITIALIZATION ---
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user_info = None


# --- LOGIN SCREEN ---
def login_screen(db):
    st.markdown(textwrap.dedent("""
    <div class="hero-header">
        <h1>🎓 PS Cadet College Karachi Exam Portal</h1>
        <p>Pakistan Steel Cadet College Karachi — Centralized Marks Management & Visual Analytics</p>
    </div>
    """), unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.container(border=True):
            st.markdown("### 🔐 Staff Authentication")
            st.caption("Sign in with your registered college email address to access portal controls.")
            
            with st.form("login_form"):
                email_input = st.text_input("Registered Email Address", placeholder="e.g. teacher@pscc.edu.pk")
                submit_button = st.form_submit_button("🔑 Login to Portal", use_container_width=True, type="primary")

                if submit_button:
                    if email_input and email_input.strip():
                        staff_df = db["Staff_Directory"]
                        user_match = staff_df[staff_df['Email'].astype(str).str.strip().str.lower() == email_input.strip().lower()]

                        if not user_match.empty:
                            st.session_state.logged_in = True
                            st.session_state.user_info = user_match.iloc[0].to_dict()
                            st.success("Authentication successful! Loading dashboard...")
                            st.rerun()
                        else:
                            st.error("❌ Email address not found in Staff Directory. Please contact the Administrator.")
                    else:
                        st.warning("⚠️ Please enter your registered email address.")


# --- HELPER LOGIC: SAFE MERGE ENGINE ---
def merge_marks_and_students(marks_df: pd.DataFrame, students_df: pd.DataFrame) -> pd.DataFrame:
    """Safely merges marks_df with students_df without producing duplicate _x and _y column suffix conflicts."""
    if marks_df.empty or students_df.empty:
        return pd.DataFrame()

    m_df = marks_df.copy()
    s_df = students_df.copy()

    # Determine join key
    if "Kit_No" in m_df.columns and "Kit_No" in s_df.columns:
        join_col = "Kit_No"
    elif "Student_ID" in m_df.columns and "Student_ID" in s_df.columns:
        join_col = "Student_ID"
    elif "Kit_No" in s_df.columns and "Student_ID" in m_df.columns:
        m_df["Kit_No"] = m_df["Student_ID"]
        join_col = "Kit_No"
    elif "Student_ID" in s_df.columns and "Kit_No" in m_df.columns:
        m_df["Student_ID"] = m_df["Kit_No"]
        join_col = "Student_ID"
    else:
        return pd.DataFrame()

    # Remove secondary ID/Group columns from m_df if present in s_df to avoid _x / _y renaming conflicts
    for col in ["Student_ID", "Kit_No", "Group", "Stream"]:
        if col != join_col and col in s_df.columns and col in m_df.columns:
            m_df = m_df.drop(columns=[col])

    merged = pd.merge(m_df, s_df, on=join_col, how="inner")

    # Ensure both Kit_No and Student_ID exist cleanly in merged dataframe
    if "Kit_No" not in merged.columns and "Student_ID" in merged.columns:
        merged["Kit_No"] = merged["Student_ID"]
    if "Student_ID" not in merged.columns and "Kit_No" in merged.columns:
        merged["Student_ID"] = merged["Kit_No"]

    return merged


# --- DATABASE MARKS SUBMISSION ---
def save_marks_to_gsheets(edited_df: pd.DataFrame, exam_name: str, subject: str, db: dict) -> int:
    """Formats edited dataframe and appends valid records to Google Sheets Marks_Log."""
    client = connect_to_gsheets()
    sheet = client.open("PS Cadet College - Master Examination Database").worksheet("Marks_Log")

    # Match Exam_ID from exam_scheme or Grading_System
    exam_scheme = db.get("exam_scheme", pd.DataFrame())
    grading_df = db.get("Grading_System", pd.DataFrame())

    exam_id = exam_name
    if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns and "Exam_ID" in exam_scheme.columns:
        match = exam_scheme.loc[exam_scheme["Exam_Name"] == exam_name, "Exam_ID"]
        if not match.empty:
            exam_id = match.values[0]
    elif not grading_df.empty and "Exam_Name" in grading_df.columns and "Exam_ID" in grading_df.columns:
        match = grading_df.loc[grading_df["Exam_Name"] == exam_name, "Exam_ID"]
        if not match.empty:
            exam_id = match.values[0]

    records_to_add = []
    assert isinstance(edited_df, pd.DataFrame)
    id_col = "Kit_No" if "Kit_No" in edited_df.columns else "Student_ID"

    for _, row in edited_df.iterrows():
        marks_val = str(row['Marks_Obtained']).strip()
        if marks_val != "" and pd.notna(row['Marks_Obtained']):
            submission_id = str(uuid.uuid4())[:8]
            student_code = str(row[id_col]).strip()
            records_to_add.append([
                submission_id,
                student_code,
                str(exam_id).strip(),
                str(subject).strip(),
                marks_val
            ])

    if records_to_add:
        sheet.append_rows(records_to_add)
        st.cache_data.clear()  # Purge cache to reflect updates instantly
        return len(records_to_add)
    return 0


# --- MAIN DASHBOARD ---
def main_dashboard(db):
    user_info = st.session_state.user_info or {}
    user_name = user_info.get('Full_Name') or user_info.get('Name') or 'Staff Member'
    user_role = user_info.get('Role') or user_info.get('Responsibility') or 'Teacher'
    class_incharge = user_info.get('Class_Incharge_Of', 'None')

    # Sidebar Profile Card
    st.sidebar.markdown(f"## 👤 {user_name}")
    st.sidebar.markdown(f"**Role:** `{user_role}`")
    if class_incharge and class_incharge.lower() != 'none':
        st.sidebar.markdown(f"**Class Incharge:** `{class_incharge}`")

    col_logout, col_refresh = st.sidebar.columns(2)
    with col_logout:
        if st.button("🚪 Logout", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.user_info = None
            st.rerun()
    with col_refresh:
        if st.button("🔄 Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    st.sidebar.divider()

    # Permissions calculation
    perm = get_staff_permissions(user_info, db)

    # Hero Banner
    st.markdown(textwrap.dedent(f"""
    <div class="hero-header">
        <h1>📚 PS Cadet College Karachi Exam Portal</h1>
        <p>Welcome back, <strong>{user_name}</strong> | Role: {user_role}</p>
    </div>
    """), unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs([
        "📊 Examination Analytics",
        "✍️ Marks Data Entry",
        "📋 Result Reports & Cadet Cards"
    ])

    # ==========================================
    # TAB 1: GLOBAL & SECTION ANALYTICS
    # ==========================================
    with tab1:
        st.subheader("📊 Global Examination Analytics & Merit Grid")

        if db["Marks_Log"].empty:
            st.info("ℹ️ No examination marks have been logged into the system yet. Use the Data Entry tab to start.")
        else:
            marks_df = db["Marks_Log"].copy()
            marks_df["Marks_Obtained"] = pd.to_numeric(marks_df["Marks_Obtained"], errors="coerce")
            students_df = db["Students"].copy()

            # Merge marks with student profile info
            merged_df = merge_marks_and_students(marks_df, students_df)

            with st.container(border=True):
                st.markdown("#### 🎯 Filter Results by Class & Exam")
                f_col1, f_col2, f_col3 = st.columns(3)

                with f_col1:
                    available_grades = sorted(merged_df["Grade"].dropna().unique())
                    dash_grade = st.selectbox("Select Grade", available_grades, key="dash_grade")

                with f_col2:
                    available_sections = sorted(merged_df[merged_df["Grade"] == dash_grade]["Section"].dropna().unique())
                    dash_section = st.selectbox("Select Section", available_sections, key="dash_section")

                with f_col3:
                    grading_df = db.get("Grading_System", pd.DataFrame())
                    exam_options = ["All Exams"]
                    if not grading_df.empty and "Exam_Name" in grading_df.columns:
                        exam_options += sorted(grading_df["Exam_Name"].dropna().unique().tolist())
                    dash_exam = st.selectbox("Select Examination", exam_options, key="dash_exam")

            # Apply Section Filters
            section_data = merged_df[(merged_df["Grade"] == dash_grade) & (merged_df["Section"] == dash_section)].copy()

            if dash_exam != "All Exams" and not grading_df.empty:
                exam_ids = grading_df.loc[grading_df["Exam_Name"] == dash_exam, "Exam_ID"].tolist()
                section_data = section_data[
                    (section_data["Exam_ID"].isin(exam_ids)) | (section_data["Exam_ID"] == dash_exam)
                ]

            if section_data.empty:
                st.warning(f"⚠️ No marks data found for Grade {dash_grade}-{dash_section} matching your filter criteria.")
            else:
                # Map Max_Marks for percentage calculation
                exam_max_map = {}
                if not grading_df.empty:
                    max_col = next((c for c in ["Max_Marks", "Total_Marks", "Full_Marks"] if c in grading_df.columns), None)
                    if max_col:
                        if "Exam_ID" in grading_df.columns:
                            for _, r in grading_df.iterrows():
                                exam_max_map[str(r["Exam_ID"]).strip()] = pd.to_numeric(r[max_col], errors="coerce")
                        if "Exam_Name" in grading_df.columns:
                            for _, r in grading_df.iterrows():
                                exam_max_map[str(r["Exam_Name"]).strip()] = pd.to_numeric(r[max_col], errors="coerce")

                if "Exam_ID" in section_data.columns and exam_max_map:
                    section_data["Max_Marks"] = section_data["Exam_ID"].astype(str).str.strip().map(exam_max_map)
                elif "Exam_Name" in section_data.columns and exam_max_map:
                    section_data["Max_Marks"] = section_data["Exam_Name"].astype(str).str.strip().map(exam_max_map)
                else:
                    section_data["Max_Marks"] = None

                section_data["Max_Marks"] = pd.to_numeric(section_data["Max_Marks"], errors="coerce").fillna(100.0)
                section_data["Max_Marks"] = section_data["Max_Marks"].replace(0, 100.0)

                valid_data = section_data[section_data["Marks_Obtained"].notna()].copy()

                # --- OVERALL METRICS CARDS ---
                total_obtained = valid_data["Marks_Obtained"].sum()
                total_max = valid_data["Max_Marks"].sum()
                overall_avg_pct = (total_obtained / total_max * 100.0) if total_max > 0 else 0.0
                grade_info = calculate_grade_info(overall_avg_pct, grading_df)

                # Student Totals & Rankings
                student_totals = valid_data.groupby(['Student_ID', 'Name']).agg(
                    Total_Obtained=('Marks_Obtained', 'sum'),
                    Total_Max=('Max_Marks', 'sum')
                ).reset_index()
                student_totals["Percentage"] = (student_totals["Total_Obtained"] / student_totals["Total_Max"] * 100.0).round(2)
                student_totals["Rank"] = student_totals["Percentage"].rank(ascending=False, method="min").astype(int)
                student_totals = student_totals.sort_values(by="Rank")

                pass_count = sum(student_totals["Percentage"] >= 40)
                pass_rate = (pass_count / len(student_totals) * 100.0) if len(student_totals) > 0 else 0.0

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Total Cadets Assessed", f"{len(student_totals)} Cadets")
                m2.metric("Class Average Marks (%)", f"{overall_avg_pct:.2f}%")
                m3.metric("Class Letter Grade", grade_info["grade"], delta=grade_info["remarks"])
                m4.metric("Class Pass Rate", f"{pass_rate:.1f}%")

                st.divider()

                # --- TOP 3 & BOTTOM 3 STANDINGS ---
                col_top, col_bottom = st.columns(2)
                with col_top:
                    with st.container(border=True):
                        st.markdown("#### 🏆 Top 3 Merit Rankers")
                        top_3 = student_totals.head(3).copy()
                        top_3["Grade"] = top_3["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["grade"])
                        st.dataframe(
                            top_3[["Rank", "Student_ID", "Name", "Total_Obtained", "Percentage", "Grade"]],
                            use_container_width=True,
                            hide_index=True
                        )

                with col_bottom:
                    with st.container(border=True):
                        st.markdown("#### ⚠️ Academic Support Needed (Bottom 3)")
                        bottom_3 = student_totals.tail(3).sort_values(by="Rank", ascending=False).copy()
                        bottom_3["Grade"] = bottom_3["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["grade"])
                        st.dataframe(
                            bottom_3[["Rank", "Student_ID", "Name", "Total_Obtained", "Percentage", "Grade"]],
                            use_container_width=True,
                            hide_index=True
                        )

                st.divider()

                # --- SUBJECT WISE SEABORN BAR CHART ---
                with st.container(border=True):
                    st.markdown("#### 📊 Subject-Wise Average Performance")
                    subj_perf = valid_data.groupby("Subject").agg(
                        Avg_Obtained=('Marks_Obtained', 'mean'),
                        Avg_Max=('Max_Marks', 'mean')
                    ).reset_index()
                    subj_perf["Avg_Percentage"] = (subj_perf["Avg_Obtained"] / subj_perf["Avg_Max"] * 100.0).round(2)

                    fig, ax = plt.subplots(figsize=(10, 4), dpi=150)
                    fig.patch.set_facecolor('#ffffff')
                    ax.set_facecolor('#f8fafc')

                    sns.barplot(
                        data=subj_perf,
                        x="Subject",
                        y="Avg_Percentage",
                        hue="Subject",
                        palette="Blues_r",
                        legend=False,
                        ax=ax
                    )
                    ax.set_ylabel("Average Score (%)", fontsize=10, fontweight='bold', color='#1e293b')
                    ax.set_xlabel("Subject", fontsize=10, fontweight='bold', color='#1e293b')
                    ax.set_ylim(0, 105)
                    ax.grid(axis='y', linestyle='--', alpha=0.5, color='#cbd5e1')
                    plt.xticks(rotation=20, ha="right", fontsize=9, fontweight='bold')

                    for p in ax.patches:
                        h = p.get_height()
                        if h > 0:
                            ax.annotate(f"{h:.1f}%", (p.get_x() + p.get_width() / 2., h / 2),
                                        ha='center', va='center', color='white', fontweight='bold', fontsize=9)

                    st.pyplot(fig)

                st.divider()

                # --- COMPREHENSIVE RESULT GRID WITH MARKS, % & GRADES ---
                with st.container(border=True):
                    st.markdown("#### 📋 Comprehensive Merit Master Sheet")

                    pivot_table = valid_data.pivot_table(
                        index=["Student_ID", "Name"],
                        columns="Subject",
                        values="Marks_Obtained",
                        aggfunc="sum"
                    ).reset_index()

                    subject_cols = [c for c in pivot_table.columns if c not in ["Student_ID", "Name"]]
                    pivot_table["Total Score"] = pivot_table[subject_cols].sum(axis=1)

                    # Merge max score and rank
                    pivot_table = pd.merge(pivot_table, student_totals[["Student_ID", "Total_Max", "Percentage", "Rank"]], on="Student_ID")
                    pivot_table["Overall Grade"] = pivot_table["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["grade"])
                    pivot_table["Status"] = pivot_table["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["status"])

                    pivot_table = pivot_table.sort_values(by="Rank").reset_index(drop=True)

                    display_cols = ["Rank", "Student_ID", "Name"] + subject_cols + ["Total Score", "Total_Max", "Percentage", "Overall Grade", "Status"]
                    st.dataframe(pivot_table[display_cols], use_container_width=True, hide_index=True)

    # ==========================================
    # TAB 2: MARKS DATA ENTRY PORTAL
    # ==========================================
    with tab2:
        st.subheader("✍️ Marks Data Entry Portal")

        # Check Teacher Specific View scoping
        is_admin = perm["is_admin"]
        assigned_grades = perm["assigned_grades"]
        assigned_sections = perm["assigned_sections"]
        assigned_subjects = perm["assigned_subjects"]

        if not is_admin and not assigned_grades:
            st.warning("""
            ⚠️ **No active teaching assignments found for your account.**
            - If you are a Subject Teacher or Class Incharge, please ask the Examination Incharge to map your assignments in `Teaching_Assignments`.
            - If you are an Administrator, please ensure your role is set to `In-charge Examination` in `Staff_Directory`.
            """)
        else:
            if is_admin:
                st.info("🔓 **Administrator Mode:** You have full access to enter marks for all grades, sections, and subjects.")
                avail_grades = sorted(db["Students"]["Grade"].dropna().unique().tolist())
            else:
                st.success(f"🔒 **Teacher Scoped Mode:** Showing only your assigned classes: `{', '.join(assigned_grades)}`")
                avail_grades = assigned_grades

            # --- SELECTION CONTROLS OUTSIDE ST.FORM FOR INSTANT REAL-TIME RE-RENDERING ---
            with st.container(border=True):
                st.markdown("#### Enter Examination & Class Selection")
                c1, c2 = st.columns(2)

                with c1:
                    exam_scheme = db.get("exam_scheme", pd.DataFrame())
                    grading_df = db.get("Grading_System", pd.DataFrame())

                    exam_opts = []
                    if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns:
                        exam_opts = sorted(exam_scheme["Exam_Name"].dropna().unique().tolist())
                    elif not grading_df.empty and "Exam_Name" in grading_df.columns:
                        exam_opts = sorted(grading_df["Exam_Name"].dropna().unique().tolist())
                    if not exam_opts:
                        exam_opts = ["Monthly_Aug", "First_Term"]

                    sel_exam = st.selectbox("Select Examination", exam_opts, key="entry_exam")
                    sel_grade = st.selectbox("Select Grade", avail_grades, key="entry_grade")

                with c2:
                    if is_admin:
                        sec_opts = sorted(db["Students"][db["Students"]["Grade"] == sel_grade]["Section"].dropna().unique().tolist())
                    else:
                        sec_opts = assigned_sections.get(sel_grade, [])

                    sel_section = st.selectbox("Select Section", sec_opts if sec_opts else ["A"], key="entry_section")

                    all_subjects = get_all_available_subjects(db)
                    if is_admin:
                        subj_opts = all_subjects
                    else:
                        subj_opts = assigned_subjects.get((sel_grade, sel_section), all_subjects)

                    sel_subject = st.selectbox("Select Subject", subj_opts if subj_opts else ["General"], key="entry_subject")

            st.divider()

            id_display_col = "Kit_No" if "Kit_No" in db["Students"].columns else "Student_ID"
            students_filtered = db["Students"][
                (db["Students"]["Grade"] == sel_grade) &
                (db["Students"]["Section"] == sel_section)
            ][[id_display_col, "Name"]].copy()

            if students_filtered.empty:
                st.warning("⚠️ No students registered in the selected Grade & Section.")
            else:
                with st.container(border=True):
                    st.markdown(f"#### Enter Marks for **Grade {sel_grade}-{sel_section}** (`{sel_subject}`)")
                    st.caption(f"Total Enrolled Cadets: {len(students_filtered)}")

                    # Check if existing marks already logged for this exam & subject to prefill grid
                    marks_log = db.get("Marks_Log", pd.DataFrame())
                    existing_map = {}
                    if not marks_log.empty and "Subject" in marks_log.columns:
                        exam_id = sel_exam
                        if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns and "Exam_ID" in exam_scheme.columns:
                            e_match = exam_scheme.loc[exam_scheme["Exam_Name"] == sel_exam, "Exam_ID"]
                            if not e_match.empty:
                                exam_id = e_match.values[0]
                        elif not grading_df.empty and "Exam_Name" in grading_df.columns and "Exam_ID" in grading_df.columns:
                            e_match = grading_df.loc[grading_df["Exam_Name"] == sel_exam, "Exam_ID"]
                            if not e_match.empty:
                                exam_id = e_match.values[0]

                        log_id_col = "Kit_No" if "Kit_No" in marks_log.columns else "Student_ID"

                        filtered_log = marks_log[
                            (marks_log["Subject"] == sel_subject) &
                            ((marks_log["Exam_ID"] == exam_id) | (marks_log["Exam_ID"] == sel_exam))
                        ]
                        existing_map = dict(zip(filtered_log[log_id_col].astype(str).str.strip(), filtered_log["Marks_Obtained"]))

                    students_filtered["Marks_Obtained"] = students_filtered[id_display_col].astype(str).str.strip().map(existing_map).fillna("")

                    # --- BULK FILE UPLOADER (CSV / EXCEL) BLOCK ---
                    with st.expander("📤 **Bulk Upload Marks via File (CSV / Excel)**", expanded=False):
                        st.markdown("""
                        Upload a **CSV** or **Excel (.xlsx / .xls)** file containing cadet examination marks to auto-fill the grid.
                        - **Supported ID Columns:** `Kit_No`, `Student_ID`, `Roll_No`, `Cadet_ID`
                        - **Supported Marks Columns:** `Marks_Obtained`, `Marks`, `Score`, `Obtained`
                        """)
                        u_col1, u_col2 = st.columns([2, 1])

                        with u_col2:
                            # Template CSV Download Button
                            tmpl_df = students_filtered.copy()
                            tmpl_csv = tmpl_df[[id_display_col, "Name", "Marks_Obtained"]].to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label="📥 Download Template (CSV)",
                                data=tmpl_csv,
                                file_name=f"PSCC_Template_Grade_{sel_grade}_{sel_section}_{sel_subject}.csv",
                                mime="text/csv",
                                use_container_width=True
                            )

                        with u_col1:
                            uploaded_file = st.file_uploader(
                                "Select CSV or Excel File",
                                type=["csv", "xlsx", "xls"],
                                key=f"bulk_upload_{sel_grade}_{sel_section}_{sel_subject}_{sel_exam}"
                            )

                        if uploaded_file is not None:
                            try:
                                if uploaded_file.name.endswith('.csv'):
                                    file_df = pd.read_csv(uploaded_file)
                                else:
                                    file_df = pd.read_excel(uploaded_file)

                                file_df.columns = [str(c).strip() for c in file_df.columns]

                                # Resolve ID & Marks Columns
                                f_id_col = next((c for c in [id_display_col, "Kit_No", "Student_ID", "Roll_No", "Cadet_ID", "ID"] if c in file_df.columns), None)
                                f_marks_col = next((c for c in ["Marks_Obtained", "Marks", "Score", "Obtained", "Mark"] if c in file_df.columns), None)

                                if f_id_col and f_marks_col:
                                    upload_map = dict(zip(file_df[f_id_col].astype(str).str.strip(), file_df[f_marks_col].astype(str).str.strip()))
                                    mapped_count = 0
                                    for idx, row in students_filtered.iterrows():
                                        s_id = str(row[id_display_col]).strip()
                                        if s_id in upload_map and upload_map[s_id] != "":
                                            students_filtered.at[idx, "Marks_Obtained"] = upload_map[s_id]
                                            mapped_count += 1

                                    st.success(f"✅ Successfully matched and auto-filled **{mapped_count}** cadet scores from `{uploaded_file.name}`! Review and save below.")
                                else:
                                    st.error(f"❌ Uploaded file must contain student ID (`Kit_No` or `Student_ID`) and marks (`Marks_Obtained` or `Marks`) headers.")
                            except Exception as ex:
                                st.error(f"❌ Error reading uploaded file: {ex}")

                    st.divider()

                    # Data Editor Grid tied to selected subject & grade key
                    edited_marks = st.data_editor(
                        students_filtered,
                        disabled=[id_display_col, "Name"],
                        hide_index=True,
                        use_container_width=True,
                        key=f"marks_editor_grid_{sel_grade}_{sel_section}_{sel_subject}_{sel_exam}"
                    )

                    if st.button("💾 Save & Update Examination Marks", use_container_width=True, type="primary"):
                        try:
                            records_saved = save_marks_to_gsheets(
                                edited_marks,
                                sel_exam,
                                sel_subject,
                                db
                            )
                            if records_saved > 0:
                                st.success(f"✅ Successfully recorded {records_saved} student mark entries to the Master Database for {sel_subject}!")
                                st.balloons()
                            else:
                                st.warning("⚠️ No marks entered. Please type a score before submitting.")
                        except Exception as e:
                            st.error(f"❌ Failed to submit marks to database: {e}")

    # ==========================================
    # TAB 3: REPORTS & CADET RESULT CARDS
    # ==========================================
    with tab3:
        st.subheader("📋 Examination Reports & Cadet Result Cards")

        sub_tab1, sub_tab2 = st.tabs(["📁 Class Master Reports & Export", "🎓 Individual Cadet Result Card"])

        # SUB TAB 1: BULK CLASS MASTER REPORTS
        with sub_tab1:
            if db["Marks_Log"].empty:
                st.info("ℹ️ No marks data available for report generation.")
            else:
                marks_df = db["Marks_Log"].copy()
                marks_df["Marks_Obtained"] = pd.to_numeric(marks_df["Marks_Obtained"], errors="coerce")
                students_df = db["Students"].copy()

                report_df = merge_marks_and_students(marks_df, students_df)

                with st.container(border=True):
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        r_grades = ["All"] + sorted(report_df["Grade"].dropna().unique().tolist())
                        f_grade = st.selectbox("Filter Grade", r_grades, key="r_grade")

                    with rc2:
                        if f_grade != "All":
                            r_sections = ["All"] + sorted(report_df[report_df["Grade"] == f_grade]["Section"].dropna().unique().tolist())
                        else:
                            r_sections = ["All"]
                        f_section = st.selectbox("Filter Section", r_sections, key="r_section")

                filt_report = report_df.copy()
                if f_grade != "All":
                    filt_report = filt_report[filt_report["Grade"] == f_grade]
                if f_section != "All":
                    filt_report = filt_report[filt_report["Section"] == f_section]

                with st.container(border=True):
                    st.markdown(f"**Total Transaction Records:** `{len(filt_report)}`")
                    disp_id_col = "Kit_No" if "Kit_No" in filt_report.columns else "Student_ID"
                    st.dataframe(
                        filt_report[[disp_id_col, "Name", "Grade", "Section", "Subject", "Marks_Obtained"]],
                        use_container_width=True,
                        height=280
                    )

                    d1, d2 = st.columns(2)
                    with d1:
                        csv_bytes = filt_report.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="⬇️ Export Class Master Sheet (CSV)",
                            data=csv_bytes,
                            file_name=f"PSCC_Marks_Report_Grade_{f_grade}_{f_section}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )

                    with d2:
                        excel_data = generate_excel_report(filt_report, sheet_name="Master_Log")
                        st.download_button(
                            label="📥 Export Styled Excel Report (.xlsx)",
                            data=excel_data,
                            file_name=f"PSCC_Marks_Report_Grade_{f_grade}_{f_section}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

        # SUB TAB 2: INDIVIDUAL CADET RESULT CARD & PRINTABLE EXPORT
        with sub_tab2:
            st.markdown("#### 🎓 Individual Cadet Result Card Generator")

            if db["Marks_Log"].empty:
                st.info("ℹ️ No marks records found in the database.")
            else:
                marks_df = db["Marks_Log"].copy()
                marks_df["Marks_Obtained"] = pd.to_numeric(marks_df["Marks_Obtained"], errors="coerce")
                students_df = db["Students"].copy()
                merged_full = merge_marks_and_students(marks_df, students_df)

                with st.container(border=True):
                    c_g, c_s, c_std, c_ex = st.columns(4)

                    with c_g:
                        avail_g = sorted(merged_full["Grade"].dropna().unique().tolist())
                        card_grade = st.selectbox("Grade", avail_g, key="card_g")

                    with c_s:
                        avail_s = sorted(merged_full[merged_full["Grade"] == card_grade]["Section"].dropna().unique().tolist())
                        card_section = st.selectbox("Section", avail_s, key="card_s")

                    with c_std:
                        cadet_df = students_df[(students_df["Grade"] == card_grade) & (students_df["Section"] == card_section)]
                        id_col_name = "Kit_No" if "Kit_No" in cadet_df.columns else "Student_ID"
                        if not cadet_df.empty:
                            cadet_options = [f"{row[id_col_name]} - {row['Name']}" for _, row in cadet_df.iterrows()]
                            kit_map = {f"{row[id_col_name]} - {row['Name']}": row[id_col_name] for _, row in cadet_df.iterrows()}
                        else:
                            cadet_options = ["No Cadets"]
                            kit_map = {}

                        selected_kit_option = st.selectbox("Kit No (Student ID)", cadet_options, key="card_kit_no")

                    with c_ex:
                        exam_scheme = db.get("exam_scheme", pd.DataFrame())
                        grading_df = db.get("Grading_System", pd.DataFrame())

                        ex_list = ["All Exams"]
                        if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns:
                            ex_list += sorted(exam_scheme["Exam_Name"].dropna().unique().tolist())
                        elif not grading_df.empty and "Exam_Name" in grading_df.columns:
                            ex_list += sorted(grading_df["Exam_Name"].dropna().unique().tolist())
                        card_exam = st.selectbox("Exam Term", ex_list, key="card_ex")

                if selected_kit_option and selected_kit_option != "No Cadets":
                    student_id = kit_map[selected_kit_option]
                    selected_student = cadet_df[cadet_df[id_col_name] == student_id].iloc[0]
                    card_student_name = selected_student["Name"]
                    student_group = selected_student.get("Group", selected_student.get("Stream", "General"))

                    # Filter student marks
                    s_marks = merged_full[
                        (merged_full["Student_ID"] == student_id) | (merged_full["Kit_No"] == student_id)
                    ].copy()
                    if card_exam != "All Exams":
                        exam_ids = []
                        if not exam_scheme.empty and "Exam_Name" in exam_scheme.columns:
                            exam_ids = exam_scheme.loc[exam_scheme["Exam_Name"] == card_exam, "Exam_ID"].tolist()
                        elif not grading_df.empty and "Exam_Name" in grading_df.columns:
                            exam_ids = grading_df.loc[grading_df["Exam_Name"] == card_exam, "Exam_ID"].tolist()

                        s_marks = s_marks[(s_marks["Exam_ID"].isin(exam_ids)) | (s_marks["Exam_ID"] == card_exam)]

                    if s_marks.empty:
                        st.warning(f"⚠️ No recorded marks found for Cadet **{card_student_name}** under **{card_exam}**.")
                    else:
                        # Calculate Max Marks mapping from exam_scheme or fallback
                        def calc_max(row):
                            ex_name = str(row.get("Exam_ID", "")).strip()
                            subj = str(row.get("Subject", "")).strip()
                            # Check exam_scheme
                            if not exam_scheme.empty and "Max_Marks" in exam_scheme.columns:
                                match = exam_scheme[
                                    ((exam_scheme["Exam_ID"].astype(str).str.strip() == ex_name) | (exam_scheme["Exam_Name"].astype(str).str.strip() == card_exam)) &
                                    (exam_scheme["Subject"].astype(str).str.strip() == subj)
                                ]
                                if not match.empty:
                                    val = pd.to_numeric(match["Max_Marks"].iloc[0], errors="coerce")
                                    if pd.notna(val) and val > 0:
                                        return float(val)

                            return 100.0

                        s_marks["Max_Marks"] = s_marks.apply(calc_max, axis=1)
                        s_marks["Percentage"] = (s_marks["Marks_Obtained"] / s_marks["Max_Marks"] * 100.0).round(2)
                        s_marks["Grade"] = s_marks["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["grade"])
                        s_marks["Remarks"] = s_marks["Percentage"].apply(lambda p: calculate_grade_info(p, grading_df)["remarks"])

                        total_obt = s_marks["Marks_Obtained"].sum()
                        total_max = s_marks["Max_Marks"].sum()
                        overall_pct = (total_obt / total_max * 100.0) if total_max > 0 else 0.0
                        overall_info = calculate_grade_info(overall_pct, grading_df)

                        # Compute Rank in Section
                        student_id_col = "Kit_No" if "Kit_No" in merged_full.columns else "Student_ID"
                        sec_totals = merged_full[
                            (merged_full["Grade"] == card_grade) & (merged_full["Section"] == card_section)
                        ].groupby(student_id_col)["Marks_Obtained"].sum().reset_index()
                        sec_totals["Rank"] = sec_totals["Marks_Obtained"].rank(ascending=False, method="min").astype(int)
                        cadet_rank_row = sec_totals[sec_totals[student_id_col] == student_id]
                        cadet_rank = cadet_rank_row["Rank"].values[0] if not cadet_rank_row.empty else "N/A"

                        st.divider()

                        # --- OFFICIAL REPORT CARD UI CONTAINER (STREAMLIT CONTAINER) ---
                        with st.container(border=True):
                            header_html = textwrap.dedent(f"""
                            <div style="text-align: center; border-bottom: 2px solid #1e3a8a; padding-bottom: 1rem; margin-bottom: 1.5rem;">
                                <h2 style="color: #1e3a8a; margin: 0; font-size: 1.8rem; font-family: 'Outfit', sans-serif;">PAKISTAN STEEL CADET COLLEGE</h2>
                                <h4 style="color: #475569; margin: 0.3rem 0; font-weight: 500;">OFFICIAL ACADEMIC EVALUATION & RESULT CARD</h4>
                                <p style="color: #64748b; margin: 0; font-size: 0.9rem;"><strong>Examination Term:</strong> {card_exam} | <strong>Academic Year:</strong> 2026</p>
                            </div>

                            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 1rem; background: #f8fafc; padding: 1.25rem; border-radius: 10px; border: 1px solid #e2e8f0; margin-bottom: 1.5rem;">
                                <div>
                                    <p style="margin: 0.25rem 0;"><strong>Cadet Name:</strong> {card_student_name}</p>
                                    <p style="margin: 0.25rem 0;"><strong>Kit / Cadet No:</strong> {student_id}</p>
                                    <p style="margin: 0.25rem 0;"><strong>Academic Group:</strong> {student_group}</p>
                                </div>
                                <div>
                                    <p style="margin: 0.25rem 0;"><strong>Grade & Section:</strong> Grade {card_grade} - {card_section}</p>
                                    <p style="margin: 0.25rem 0;"><strong>Merit Position (Section):</strong> #{cadet_rank} out of {len(sec_totals)} Cadets</p>
                                    <p style="margin: 0.25rem 0;"><strong>Evaluation Date:</strong> 2026-07-23</p>
                                </div>
                            </div>
                            """)
                            st.markdown(header_html, unsafe_allow_html=True)

                            # Summary Metrics
                            sum1, sum2, sum3, sum4 = st.columns(4)
                            sum1.metric("Total Marks Obtained", f"{total_obt} / {total_max}")
                            sum2.metric("Aggregate Percentage", f"{overall_pct:.2f}%")
                            sum3.metric("Final Grade", overall_info["grade"])
                            sum4.metric("Academic Status", overall_info["status"])

                            st.markdown("##### 📝 Subject Score Breakdown")
                            st.dataframe(
                                s_marks[["Subject", "Marks_Obtained", "Max_Marks", "Percentage", "Grade", "Remarks"]],
                                use_container_width=True,
                                hide_index=True
                            )

                            # Cadet vs Class Average Visual
                            st.markdown("##### 📊 Performance Comparison vs Class Average")
                            class_subj_avg = merged_full[
                                (merged_full["Grade"] == card_grade) & (merged_full["Section"] == card_section)
                            ].groupby("Subject")["Marks_Obtained"].mean().reset_index()

                            comp_df = pd.merge(s_marks[["Subject", "Marks_Obtained"]], class_subj_avg, on="Subject", suffixes=("_Cadet", "_Class_Avg"))

                            fig, ax = plt.subplots(figsize=(9, 3.5), dpi=150)
                            fig.patch.set_facecolor('#ffffff')
                            ax.set_facecolor('#f8fafc')

                            x_indices = range(len(comp_df))
                            width = 0.35

                            ax.bar([x - width/2 for x in x_indices], comp_df["Marks_Obtained_Cadet"], width, label=f"Cadet: {card_student_name}", color="#1e3a8a")
                            ax.bar([x + width/2 for x in x_indices], comp_df["Marks_Obtained_Class_Avg"], width, label="Class Average", color="#94a3b8")

                            ax.set_ylabel("Marks Obtained", fontsize=9, fontweight='bold')
                            ax.set_xticks(list(x_indices))
                            ax.set_xticklabels(comp_df["Subject"], rotation=15, ha='right', fontsize=9, fontweight='bold')
                            ax.grid(axis='y', linestyle='--', alpha=0.4, color='#cbd5e1')
                            ax.legend(frameon=True, facecolor='white', edgecolor='#cbd5e1')
                            st.pyplot(fig)

                            # Signature Footer Block
                            footer_html = textwrap.dedent("""
                            <div style="margin-top: 2rem; padding-top: 1.5rem; border-top: 2px dashed #cbd5e1; display: flex; flex-wrap: wrap; justify-content: space-between; gap: 1rem; text-align: center;">
                                <div style="flex: 1; min-width: 150px;"><br>___________________<br><strong>Class Incharge</strong></div>
                                <div style="flex: 1; min-width: 150px;"><br>___________________<br><strong>In-charge Examination</strong></div>
                                <div style="flex: 1; min-width: 150px;"><br>___________________<br><strong>Principal / Controller</strong></div>
                            </div>
                            """)
                            st.markdown(footer_html, unsafe_allow_html=True)

                        st.divider()

                        # Export Buttons for Individual Cadet Card
                        exp1, exp2 = st.columns(2)
                        with exp1:
                            cadet_excel = generate_excel_report(
                                s_marks[["Subject", "Marks_Obtained", "Max_Marks", "Percentage", "Grade", "Remarks"]],
                                sheet_name=f"Report_{student_id}"
                            )
                            st.download_button(
                                label="📥 Download Cadet Report Card (.xlsx)",
                                data=cadet_excel,
                                file_name=f"PSCC_Report_Card_{student_id}_{card_student_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

                        with exp2:
                            # Printable HTML card
                            html_content = textwrap.dedent(f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <title>Report Card - {card_student_name}</title>
                                <style>
                                    body {{ font-family: 'Inter', Arial, sans-serif; padding: 20px; color: #0f172a; }}
                                    .header {{ text-align: center; border-bottom: 2px solid #1e3a8a; padding-bottom: 10px; }}
                                    .info {{ margin: 20px 0; background: #f8fafc; padding: 15px; border-radius: 8px; border: 1px solid #e2e8f0; }}
                                    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                                    th, td {{ border: 1px solid #cbd5e1; padding: 10px; text-align: left; }}
                                    th {{ background-color: #1e3a8a; color: white; }}
                                </style>
                            </head>
                            <body>
                                <div class="header">
                                    <h2 style="color: #1e3a8a;">PAKISTAN STEEL CADET COLLEGE</h2>
                                    <h3>OFFICIAL RESULT CARD - {card_exam}</h3>
                                </div>
                                <div class="info">
                                    <p><strong>Cadet Name:</strong> {card_student_name} | <strong>ID:</strong> {student_id}</p>
                                    <p><strong>Grade & Section:</strong> {card_grade}-{card_section} | <strong>Position:</strong> #{cadet_rank}</p>
                                    <p><strong>Total Score:</strong> {total_obt}/{total_max} ({overall_pct:.2f}%) | <strong>Grade:</strong> {overall_info['grade']}</p>
                                </div>
                                <table>
                                    <tr><th>Subject</th><th>Marks</th><th>Max</th><th>Percentage</th><th>Grade</th><th>Remarks</th></tr>
                                    {"".join([f"<tr><td>{r['Subject']}</td><td>{r['Marks_Obtained']}</td><td>{r['Max_Marks']}</td><td>{r['Percentage']}%</td><td>{r['Grade']}</td><td>{r['Remarks']}</td></tr>" for _, r in s_marks.iterrows()])}
                                </table>
                                <br><br>
                                <button onclick="window.print()" style="padding:12px 24px; background:#1e3a8a; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:bold;">🖨️ Print / Save as PDF</button>
                            </body>
                            </html>
                            """)
                            st.download_button(
                                label="🖨️ Download Printable HTML / PDF Card",
                                data=html_content,
                                file_name=f"PSCC_Report_Card_{student_id}_{card_student_name}.html",
                                mime="text/html",
                                use_container_width=True
                            )


# --- MAIN APP ROUTING ---
try:
    with st.spinner("🔄 Connecting to PSCC Master Database..."):
        app_db = load_database()

    if not st.session_state.logged_in:
        login_screen(app_db)
    else:
        main_dashboard(app_db)

except Exception as e:
    st.error(f"❌ Failed to load application interface. Error details: {e}")